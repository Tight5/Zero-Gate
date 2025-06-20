"""
Integration Agent for Zero Gate ESO Platform
Handles Microsoft Graph integration with MSAL authentication, organizational relationship extraction,
email communication analysis, and Excel file processing for dashboard data
"""

import logging
import asyncio
import json
import os
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import requests
from msal import ConfidentialClientApplication
import networkx as nx
from collections import defaultdict, Counter

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("zero-gate.integration")

class IntegrationAgent:
    """
    Microsoft Graph Integration Agent with MSAL authentication
    Provides organizational data extraction, email analysis, and Excel processing
    """
    
    def __init__(self, resource_monitor=None):
        self.resource_monitor = resource_monitor
        self.graph_clients = {}
        self.token_cache = {}
        self.base_url = "https://graph.microsoft.com/v1.0"
        
        # Microsoft Graph scopes for different operations
        self.scopes = [
            "https://graph.microsoft.com/User.Read.All",
            "https://graph.microsoft.com/Directory.Read.All", 
            "https://graph.microsoft.com/Mail.Read",
            "https://graph.microsoft.com/Files.Read.All",
            "https://graph.microsoft.com/People.Read.All"
        ]
        
        logger.info("IntegrationAgent initialized with Microsoft Graph support")
    
    def get_access_token(self, tenant_id: Optional[str] = None) -> Optional[str]:
        """Get or refresh Microsoft Graph access token using MSAL"""
        try:
            # Use environment tenant if not specified
            effective_tenant_id = tenant_id or os.getenv("MICROSOFT_TENANT_ID")
            if not effective_tenant_id:
                logger.error("No tenant ID provided")
                return None
            
            # Check if we have a valid cached token
            if effective_tenant_id in self.token_cache:
                token_info = self.token_cache[effective_tenant_id]
                if datetime.now() < token_info["expires_at"] - timedelta(minutes=5):
                    return token_info["access_token"]
            
            # Create MSAL client application
            client_app = ConfidentialClientApplication(
                client_id=os.getenv("MICROSOFT_CLIENT_ID"),
                client_credential=os.getenv("MICROSOFT_CLIENT_SECRET"),
                authority=f"https://login.microsoftonline.com/{effective_tenant_id}"
            )
            
            # Acquire token using client credentials flow
            result = client_app.acquire_token_for_client(
                scopes=["https://graph.microsoft.com/.default"]
            )
            
            if result and isinstance(result, dict) and "access_token" in result:
                # Cache the token
                expires_in = result.get("expires_in", 3600)
                if isinstance(expires_in, (int, float)):
                    self.token_cache[effective_tenant_id] = {
                        "access_token": result["access_token"],
                        "expires_at": datetime.now() + timedelta(seconds=float(expires_in))
                    }
                    
                    logger.info(f"Successfully acquired access token for tenant {effective_tenant_id}")
                    return result["access_token"]
            else:
                error_desc = "Unknown error"
                if isinstance(result, dict):
                    error_desc = result.get("error_description", "Unknown error")
                logger.error(f"Failed to acquire token: {error_desc}")
                return None
                
        except Exception as e:
            logger.error(f"Error acquiring access token: {str(e)}")
            return None
    
    def make_graph_request(self, endpoint: str, method: str = "GET", 
                          data: Optional[Dict] = None, tenant_id: Optional[str] = None) -> Optional[Dict]:
        """Make authenticated request to Microsoft Graph API"""
        access_token = self.get_access_token(tenant_id)
        if not access_token:
            return None
        
        headers = {
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/json"
        }
        
        url = f"{self.base_url}/{endpoint}"
        
        try:
            if method == "GET":
                response = requests.get(url, headers=headers)
            elif method == "POST":
                response = requests.post(url, headers=headers, json=data)
            else:
                logger.error(f"Unsupported HTTP method: {method}")
                return None
            
            if response.status_code == 200:
                return response.json()
            else:
                logger.error(f"Graph API request failed: {response.status_code} - {response.text}")
                return None
                
        except Exception as e:
            logger.error(f"Error making Graph request to {endpoint}: {str(e)}")
            return None
    
    async def extract_organizational_relationships(self, tenant_id: str = None, 
                                                 user_limit: int = 100) -> Dict[str, Any]:
        """
        Extract organizational relationships from Microsoft Graph
        Returns manager/report hierarchies and collaboration networks
        """
        logger.info("Starting organizational relationship extraction")
        
        try:
            # Get all users in the organization
            users_response = self.make_graph_request(
                f"users?$select=id,displayName,mail,jobTitle,department,manager&$top={user_limit}",
                tenant_id=tenant_id
            )
            
            if not users_response or "value" not in users_response:
                return {"status": "error", "message": "Failed to retrieve users"}
            
            users = users_response["value"]
            relationships = []
            user_mapping = {}
            
            # Create user mapping for quick lookup
            for user in users:
                user_mapping[user["id"]] = {
                    "id": user["id"],
                    "name": user.get("displayName", "Unknown"),
                    "email": user.get("mail", ""),
                    "title": user.get("jobTitle", ""),
                    "department": user.get("department", "")
                }
            
            # Extract manager relationships
            for user in users:
                user_id = user["id"]
                
                # Get manager information
                manager_response = self.make_graph_request(
                    f"users/{user_id}/manager?$select=id,displayName,mail",
                    tenant_id=tenant_id
                )
                
                if manager_response and "id" in manager_response:
                    manager_id = manager_response["id"]
                    
                    relationships.append({
                        "from": manager_id,
                        "to": user_id,
                        "type": "manager",
                        "strength": 1.0,
                        "created_at": datetime.now().isoformat()
                    })
                
                # Get direct reports
                reports_response = self.make_graph_request(
                    f"users/{user_id}/directReports?$select=id,displayName,mail",
                    tenant_id=tenant_id
                )
                
                if reports_response and "value" in reports_response:
                    for report in reports_response["value"]:
                        report_id = report["id"]
                        
                        relationships.append({
                            "from": user_id,
                            "to": report_id,
                            "type": "direct_report",
                            "strength": 1.0,
                            "created_at": datetime.now().isoformat()
                        })
            
            return {
                "status": "success",
                "tenant_id": tenant_id,
                "users": user_mapping,
                "relationships": relationships,
                "total_users": len(users),
                "total_relationships": len(relationships),
                "extracted_at": datetime.now().isoformat()
            }
            
        except Exception as e:
            logger.error(f"Error extracting organizational relationships: {str(e)}")
            return {"status": "error", "message": str(e)}
    
    async def analyze_email_communication_patterns(self, tenant_id: str = None, 
                                                  user_id: str = None, 
                                                  days: int = 30) -> Dict[str, Any]:
        """
        Analyze email communication patterns to determine relationship strength
        Returns communication frequency, response patterns, and collaboration scores
        """
        logger.info(f"Analyzing email communication patterns for {days} days")
        
        try:
            # If no specific user, analyze for current authenticated user
            if not user_id:
                me_response = self.make_graph_request("me?$select=id,displayName,mail", tenant_id=tenant_id)
                if not me_response:
                    return {"status": "error", "message": "Cannot identify current user"}
                user_id = me_response["id"]
            
            # Calculate date range
            end_date = datetime.now()
            start_date = end_date - timedelta(days=days)
            
            # Get email messages with metadata
            filter_query = f"receivedDateTime ge {start_date.isoformat()} and receivedDateTime le {end_date.isoformat()}"
            
            messages_response = self.make_graph_request(
                f"users/{user_id}/messages?$filter={filter_query}&$select=sender,toRecipients,ccRecipients,receivedDateTime,isRead,hasAttachments,importance&$top=1000",
                tenant_id=tenant_id
            )
            
            if not messages_response or "value" not in messages_response:
                return {"status": "error", "message": "Failed to retrieve email messages"}
            
            messages = messages_response["value"]
            
            # Analyze communication patterns
            communication_data = defaultdict(lambda: {
                "email": "",
                "name": "",
                "message_count": 0,
                "sent_count": 0,
                "received_count": 0,
                "response_rate": 0.0,
                "avg_response_time_hours": 0.0,
                "collaboration_score": 0.0,
                "importance_high": 0,
                "has_attachments": 0
            })
            
            # Process messages
            for message in messages:
                # Sender analysis
                if message.get("sender") and message["sender"].get("emailAddress"):
                    sender_email = message["sender"]["emailAddress"]["address"]
                    sender_name = message["sender"]["emailAddress"]["name"]
                    
                    communication_data[sender_email]["email"] = sender_email
                    communication_data[sender_email]["name"] = sender_name
                    communication_data[sender_email]["received_count"] += 1
                    communication_data[sender_email]["message_count"] += 1
                    
                    # Check importance and attachments
                    if message.get("importance") == "high":
                        communication_data[sender_email]["importance_high"] += 1
                    
                    if message.get("hasAttachments"):
                        communication_data[sender_email]["has_attachments"] += 1
                
                # Recipients analysis (for sent emails)
                for recipient_list in [message.get("toRecipients", []), message.get("ccRecipients", [])]:
                    for recipient in recipient_list:
                        if recipient.get("emailAddress"):
                            recipient_email = recipient["emailAddress"]["address"]
                            recipient_name = recipient["emailAddress"]["name"]
                            
                            communication_data[recipient_email]["email"] = recipient_email
                            communication_data[recipient_email]["name"] = recipient_name
                            communication_data[recipient_email]["sent_count"] += 1
                            communication_data[recipient_email]["message_count"] += 1
            
            # Calculate collaboration scores
            total_messages = sum(data["message_count"] for data in communication_data.values())
            
            for email, data in communication_data.items():
                if total_messages > 0:
                    # Base score on message frequency
                    frequency_score = min(1.0, data["message_count"] / total_messages * 20)
                    
                    # Boost score for bidirectional communication
                    bidirectional_bonus = 0.0
                    if data["sent_count"] > 0 and data["received_count"] > 0:
                        bidirectional_bonus = 0.3
                    
                    # Boost for high importance messages
                    importance_bonus = min(0.2, data["importance_high"] / max(1, data["message_count"]) * 0.5)
                    
                    # Boost for attachments (indicates collaboration)
                    attachment_bonus = min(0.2, data["has_attachments"] / max(1, data["message_count"]) * 0.4)
                    
                    data["collaboration_score"] = min(1.0, 
                        frequency_score + bidirectional_bonus + importance_bonus + attachment_bonus
                    )
                    
                    # Calculate response rate (simplified)
                    if data["received_count"] > 0:
                        data["response_rate"] = min(1.0, data["sent_count"] / data["received_count"])
            
            # Convert to regular dict and sort by collaboration score
            result_data = dict(communication_data)
            sorted_contacts = sorted(
                result_data.items(), 
                key=lambda x: x[1]["collaboration_score"], 
                reverse=True
            )
            
            return {
                "status": "success",
                "tenant_id": tenant_id,
                "user_id": user_id,
                "analysis_period_days": days,
                "total_contacts": len(result_data),
                "total_messages_analyzed": len(messages),
                "communication_patterns": dict(sorted_contacts[:50]),  # Top 50 contacts
                "top_collaborators": [
                    {
                        "email": email,
                        "name": data["name"],
                        "score": data["collaboration_score"],
                        "message_count": data["message_count"]
                    }
                    for email, data in sorted_contacts[:10]  # Top 10
                ],
                "analyzed_at": datetime.now().isoformat()
            }
            
        except Exception as e:
            logger.error(f"Error analyzing communication patterns: {str(e)}")
            return {"status": "error", "message": str(e)}
    
    async def process_excel_file_for_dashboard(self, file_path: str = None, 
                                             file_content: bytes = None,
                                             tenant_id: str = None) -> Dict[str, Any]:
        """
        Process Excel file to extract dashboard-relevant data
        Supports both local files and file content from uploads
        """
        logger.info("Processing Excel file for dashboard data extraction")
        
        try:
            # Load workbook
            if file_path:
                workbook = load_workbook(file_path, data_only=True)
            elif file_content:
                from io import BytesIO
                workbook = load_workbook(BytesIO(file_content), data_only=True)
            else:
                return {"status": "error", "message": "No file path or content provided"}
            
            dashboard_data = {
                "worksheets": [],
                "kpi_data": {},
                "charts_data": [],
                "tables": [],
                "sponsor_data": [],
                "grant_data": [],
                "financial_data": []
            }
            
            # Process each worksheet
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                sheet_data = {
                    "name": sheet_name,
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                    "data_extracted": []
                }
                
                # Convert worksheet to pandas DataFrame for easier processing
                data = []
                for row in worksheet.iter_rows(values_only=True):
                    data.append(row)
                
                if data:
                    df = pd.DataFrame(data[1:], columns=data[0])  # First row as headers
                    
                    # Detect and extract KPI-like data
                    kpi_patterns = self._extract_kpi_data(df, sheet_name)
                    if kpi_patterns:
                        dashboard_data["kpi_data"][sheet_name] = kpi_patterns
                    
                    # Detect sponsor data
                    sponsor_patterns = self._extract_sponsor_data(df, sheet_name)
                    if sponsor_patterns:
                        dashboard_data["sponsor_data"].extend(sponsor_patterns)
                    
                    # Detect grant data
                    grant_patterns = self._extract_grant_data(df, sheet_name)
                    if grant_patterns:
                        dashboard_data["grant_data"].extend(grant_patterns)
                    
                    # Detect financial data
                    financial_patterns = self._extract_financial_data(df, sheet_name)
                    if financial_patterns:
                        dashboard_data["financial_data"].extend(financial_patterns)
                    
                    # Store processed data summary
                    sheet_data["data_extracted"] = {
                        "rows": len(df),
                        "columns": len(df.columns),
                        "column_names": list(df.columns),
                        "data_types": df.dtypes.to_dict()
                    }
                
                dashboard_data["worksheets"].append(sheet_data)
            
            return {
                "status": "success",
                "tenant_id": tenant_id,
                "file_processed": True,
                "dashboard_data": dashboard_data,
                "processed_at": datetime.now().isoformat(),
                "summary": {
                    "total_worksheets": len(dashboard_data["worksheets"]),
                    "kpi_sheets": len(dashboard_data["kpi_data"]),
                    "sponsor_records": len(dashboard_data["sponsor_data"]),
                    "grant_records": len(dashboard_data["grant_data"]),
                    "financial_records": len(dashboard_data["financial_data"])
                }
            }
            
        except Exception as e:
            logger.error(f"Error processing Excel file: {str(e)}")
            return {"status": "error", "message": str(e)}
    
    def _extract_kpi_data(self, df: pd.DataFrame, sheet_name: str) -> Dict[str, Any]:
        """Extract KPI-like data from DataFrame"""
        kpi_data = {}
        
        # Look for common KPI patterns
        kpi_keywords = [
            "total", "count", "sum", "average", "percentage", "rate", "score",
            "sponsors", "grants", "funding", "applications", "success", "revenue"
        ]
        
        for column in df.columns:
            if any(keyword in str(column).lower() for keyword in kpi_keywords):
                # Try to extract numeric values
                numeric_data = pd.to_numeric(df[column], errors='coerce').dropna()
                if not numeric_data.empty:
                    kpi_data[column] = {
                        "value": numeric_data.iloc[-1] if len(numeric_data) == 1 else numeric_data.sum(),
                        "data_type": "numeric",
                        "source_sheet": sheet_name
                    }
        
        return kpi_data
    
    def _extract_sponsor_data(self, df: pd.DataFrame, sheet_name: str) -> List[Dict[str, Any]]:
        """Extract sponsor-related data from DataFrame"""
        sponsor_data = []
        
        # Look for sponsor-related columns
        sponsor_columns = ["sponsor", "organization", "company", "funder", "donor"]
        name_columns = ["name", "contact", "person", "representative"]
        email_columns = ["email", "contact_email", "e-mail"]
        
        sponsor_col = None
        name_col = None
        email_col = None
        
        # Find relevant columns
        for col in df.columns:
            col_lower = str(col).lower()
            if any(keyword in col_lower for keyword in sponsor_columns):
                sponsor_col = col
            elif any(keyword in col_lower for keyword in name_columns):
                name_col = col
            elif any(keyword in col_lower for keyword in email_columns):
                email_col = col
        
        # Extract data if sponsor column found
        if sponsor_col:
            for _, row in df.iterrows():
                if pd.notna(row[sponsor_col]):
                    sponsor_record = {
                        "organization": str(row[sponsor_col]),
                        "contact_name": str(row[name_col]) if name_col and pd.notna(row[name_col]) else "",
                        "email": str(row[email_col]) if email_col and pd.notna(row[email_col]) else "",
                        "source_sheet": sheet_name,
                        "extracted_at": datetime.now().isoformat()
                    }
                    sponsor_data.append(sponsor_record)
        
        return sponsor_data
    
    def _extract_grant_data(self, df: pd.DataFrame, sheet_name: str) -> List[Dict[str, Any]]:
        """Extract grant-related data from DataFrame"""
        grant_data = []
        
        # Look for grant-related columns
        grant_columns = ["grant", "funding", "award", "application", "proposal"]
        amount_columns = ["amount", "value", "funding", "budget", "total"]
        date_columns = ["date", "deadline", "due", "submission"]
        
        grant_col = None
        amount_col = None
        date_col = None
        
        # Find relevant columns
        for col in df.columns:
            col_lower = str(col).lower()
            if any(keyword in col_lower for keyword in grant_columns):
                grant_col = col
            elif any(keyword in col_lower for keyword in amount_columns):
                amount_col = col
            elif any(keyword in col_lower for keyword in date_columns):
                date_col = col
        
        # Extract data if grant column found
        if grant_col:
            for _, row in df.iterrows():
                if pd.notna(row[grant_col]):
                    grant_record = {
                        "title": str(row[grant_col]),
                        "amount": str(row[amount_col]) if amount_col and pd.notna(row[amount_col]) else "",
                        "deadline": str(row[date_col]) if date_col and pd.notna(row[date_col]) else "",
                        "source_sheet": sheet_name,
                        "extracted_at": datetime.now().isoformat()
                    }
                    grant_data.append(grant_record)
        
        return grant_data
    
    def _extract_financial_data(self, df: pd.DataFrame, sheet_name: str) -> List[Dict[str, Any]]:
        """Extract financial data from DataFrame"""
        financial_data = []
        
        # Look for financial columns
        financial_keywords = ["revenue", "income", "expense", "cost", "budget", "profit", "loss"]
        
        for column in df.columns:
            if any(keyword in str(column).lower() for keyword in financial_keywords):
                # Extract numeric financial data
                numeric_data = pd.to_numeric(df[column], errors='coerce').dropna()
                if not numeric_data.empty:
                    financial_record = {
                        "category": column,
                        "total": float(numeric_data.sum()),
                        "average": float(numeric_data.mean()),
                        "count": int(len(numeric_data)),
                        "source_sheet": sheet_name,
                        "extracted_at": datetime.now().isoformat()
                    }
                    financial_data.append(financial_record)
        
        return financial_data
    
    async def get_connection_status(self, tenant_id: str = None) -> Dict[str, Any]:
        """Check Microsoft Graph connection status and permissions"""
        try:
            access_token = self.get_access_token(tenant_id)
            if not access_token:
                return {
                    "status": "disconnected",
                    "message": "Unable to acquire access token",
                    "authenticated": False
                }
            
            # Test connection with a simple request
            me_response = self.make_graph_request("me?$select=id,displayName,mail", tenant_id=tenant_id)
            
            if me_response:
                return {
                    "status": "connected",
                    "authenticated": True,
                    "user": {
                        "id": me_response.get("id"),
                        "name": me_response.get("displayName"),
                        "email": me_response.get("mail")
                    },
                    "tenant_id": tenant_id,
                    "scopes_available": self.scopes,
                    "last_checked": datetime.now().isoformat()
                }
            else:
                return {
                    "status": "error",
                    "message": "Failed to connect to Microsoft Graph",
                    "authenticated": False
                }
                
        except Exception as e:
            logger.error(f"Error checking connection status: {str(e)}")
            return {
                "status": "error",
                "message": str(e),
                "authenticated": False
            }

# Global integration agent instance
integration_agent = None

def get_integration_agent():
    """Get the global integration agent instance"""
    global integration_agent
    if integration_agent is None:
        integration_agent = IntegrationAgent()
    return integration_agent